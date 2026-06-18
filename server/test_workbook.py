"""Generate a detailed test-case Excel workbook from a TestPlan + TestSuites.

Produces a multi-sheet .xlsx that a QA/dev team can use directly:

  • Summary       — strategy, coverage target, framework, counts
  • Smoke         — critical-path must-pass checks (fast confidence)
  • Regression    — full functional coverage (every scenario)
  • End-to-End    — routed flows / integration / manual journeys

Each test row has an ID, area, scenario, pre-conditions, numbered steps,
expected result, priority, type, and an empty Status column for the tester.
The workbook is built in-memory and returned as bytes (no temp files).
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_HEADERS = ["Test ID", "Area / Component", "Category", "Type", "Priority",
            "Scenario", "Pre-conditions", "Steps", "Expected Result", "Status"]

_DHL_RED = "D40511"
_HDR_FILL = PatternFill("solid", fgColor=_DHL_RED)
_HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
_WRAP = Alignment(wrap_text=True, vertical="top")
_THIN = Side(style="thin", color="DDDDDD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_COL_W = [12, 26, 12, 13, 9, 34, 28, 40, 40, 10]


def _steps_for(scenario: str, area: str) -> str:
    """Synthesize numbered, tester-usable steps from a scenario description."""
    s = scenario.strip().rstrip(".")
    return (f"1. Mount/navigate to {area}.\n"
            f"2. {s[0].upper() + s[1:]}.\n"
            f"3. Observe the rendered output and any callbacks/network calls.")


def _expected_for(scenario: str) -> str:
    s = scenario.strip().rstrip(".")
    return f"The component behaves as described: {s}. No console errors; UI matches the migrated design."


def _precond_for(area: str, type_: str) -> str:
    base = f"The migrated React app is built and running; {area} is importable."
    if type_ == "integration":
        base += " Backend/service mocks are configured (vi.fn / MSW)."
    return base


def _style_sheet(ws, rows: list[list[Any]]) -> None:
    ws.append(_HEADERS)
    for c in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.border = _HDR_FILL, _HDR_FONT, _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = _COL_W[c - 1]
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(_HEADERS)):
        for cell in row:
            cell.alignment, cell.border = _WRAP, _BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}{ws.max_row}"


def _matrix_rows(test_plan: dict) -> tuple[list, list, list]:
    """Split the TestPlan matrix into (smoke, regression, e2e) row lists."""
    smoke, regression, e2e = [], [], []
    sn = rn = en = 0
    for entry in test_plan.get("matrix", []):
        area = entry.get("target", "Component")
        type_ = entry.get("type", "unit")
        prio = entry.get("priority", "medium")
        scenarios = entry.get("scenarios", []) or []
        for i, sc in enumerate(scenarios):
            pre = _precond_for(area, type_)
            steps = _steps_for(sc, area)
            exp = _expected_for(sc)
            # Regression = every scenario.
            rn += 1
            regression.append([f"REG-{rn:03d}", area, "Regression", type_, prio,
                               sc, pre, steps, exp, ""])
            # Smoke = the first scenario of each high-priority target (must-pass).
            if i == 0 and prio == "high":
                sn += 1
                smoke.append([f"SMK-{sn:03d}", area, "Smoke", type_, prio,
                              sc, pre, steps, exp, ""])
            # E2E = integration-type scenarios (routed / service flows).
            if type_ == "integration":
                en += 1
                e2e.append([f"E2E-{en:03d}", area, "End-to-End", type_, prio,
                            sc, pre, steps, exp, ""])
    # E2E also includes any manual journey checklist items.
    for item in test_plan.get("manual_checklist", []) or []:
        en += 1
        e2e.append([f"E2E-{en:03d}", "Full app", "End-to-End", "manual", "high",
                    item, "App deployed to a test environment.",
                    f"1. {item}.", "Behaviour matches the original Angular app.", ""])
    return smoke, regression, e2e


def build_test_workbook(test_plan: dict, suites: list[dict] | None = None) -> bytes:
    """Build the categorized test-case workbook and return it as .xlsx bytes."""
    test_plan = test_plan or {}
    suites = suites or []
    smoke, regression, e2e = _matrix_rows(test_plan)

    # Append concrete generated unit tests (from the suites) to Regression.
    rn = len(regression)
    for suite in suites:
        for tc in suite.get("tests", []):
            rn += 1
            covers = ", ".join(tc.get("covers", []) or [])
            regression.append([
                f"REG-{rn:03d}", tc.get("file_path", "—"), "Regression",
                tc.get("type", "unit"), "medium", tc.get("name", "generated test"),
                "Generated Vitest test compiles and the component renders.",
                "1. Run `npx vitest run`.\n2. Execute this test case.",
                f"Test passes. Covers: {covers or 'component behaviour'}.", ""])

    wb = Workbook()
    # Summary sheet
    ws0 = wb.active
    ws0.title = "Summary"
    ws0.append(["NextGen Refactor — Test Case Document"])
    ws0["A1"].font = Font(bold=True, size=14, color=_DHL_RED)
    ws0.append([])
    meta = [
        ("Strategy", test_plan.get("strategy_summary", "—")),
        ("Framework", test_plan.get("framework", "vitest")),
        ("Coverage target", f"{test_plan.get('coverage_target_pct', 0)}%"),
        ("Mocking strategy", test_plan.get("mocking_strategy", "—")),
        ("Smoke tests", len(smoke)),
        ("Regression tests", len(regression)),
        ("End-to-End tests", len(e2e)),
        ("Total tests", len(smoke) + len(regression) + len(e2e)),
    ]
    for k, v in meta:
        ws0.append([k, v])
        ws0.cell(row=ws0.max_row, column=1).font = Font(bold=True)
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 90
    for row in ws0.iter_rows(min_row=3, max_row=ws0.max_row, min_col=2, max_col=2):
        row[0].alignment = _WRAP

    _style_sheet(wb.create_sheet("Smoke"), smoke)
    _style_sheet(wb.create_sheet("Regression"), regression)
    _style_sheet(wb.create_sheet("End-to-End"), e2e)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
