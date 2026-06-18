"""Pytest configuration — marker registration and XLSX test report generation.

Every test run writes ``test_report.xlsx`` to the project root.  The report is
channelised into one sheet per test category:

    Summary | Smoke | Regression | E2E Positive | E2E Negative | API | Integration

Category is inferred from the explicit pytest marker on the test, falling back
to the test module filename when no marker is present.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pytest

# ---------------------------------------------------------------------------
# Marker definitions
# ---------------------------------------------------------------------------

_MARKERS = [
    ("smoke",          "Quick sanity checks — pure Python, no ADK required"),
    ("regression",     "Regression guards against previously reported bugs"),
    ("e2e_positive",   "End-to-end happy-path scenarios through the full pipeline"),
    ("e2e_negative",   "End-to-end error, rejection, and edge-case scenarios"),
    ("api",            "HTTP API surface tests (FastAPI / auth)"),
    ("integration",    "Full ADK pipeline integration tests (require google-adk)"),
    ("transformation", "Transformation-process validation tests (parser + validator)"),
]

_CATEGORY_ORDER = [m[0] for m in _MARKERS]

_FILENAME_TO_CATEGORY: dict[str, str] = {
    "test_smoke":                    "smoke",
    "test_regression":               "regression",
    "test_e2e_positive":             "e2e_positive",
    "test_e2e_negative":             "e2e_negative",
    "test_v3_api":                   "api",
    "test_v3_auth":                  "api",
    "test_v3_phase1":                "integration",
    "test_v3_phase2":                "integration",
    "test_v3_phase3":                "integration",
    "test_v3_phase4":                "integration",
    "test_v3_service":               "integration",
    "test_v3_workflow":              "integration",
    "test_v3_ingestion":             "integration",
    "test_transformation_validation": "transformation",
}

# ---------------------------------------------------------------------------
# Session-level state (safe: pytest runs single-threaded)
# ---------------------------------------------------------------------------

_results: list[dict[str, Any]] = []
_nodeid_to_category: dict[str, str] = {}

# ---------------------------------------------------------------------------
# Hooks
# ---------------------------------------------------------------------------


def pytest_configure(config: pytest.Config) -> None:
    for name, desc in _MARKERS:
        config.addinivalue_line("markers", f"{name}: {desc}")


def pytest_collection_modifyitems(items: list[pytest.Item]) -> None:
    for item in items:
        _nodeid_to_category[item.nodeid] = _infer_category(item)


def pytest_runtest_logreport(report: pytest.TestReport) -> None:
    if report.when == "call":
        if report.passed:
            status = "PASSED"
        elif report.failed:
            status = "FAILED"
        elif report.skipped:
            status = "SKIPPED"
        else:
            return
        _append_result(report, status)
    elif report.when == "setup":
        if report.skipped:
            _append_result(report, "SKIPPED")
        elif report.failed:
            _append_result(report, "ERROR")


def pytest_sessionfinish(session: pytest.Session, exitstatus: int) -> None:
    if not _results:
        return
    try:
        out = Path(__file__).parent / "test_report.xlsx"
        _write_xlsx(_results, out)
        print(f"\n[REPORT] Test report written: {out}")
    except Exception as exc:  # never break a test run over report generation
        import traceback
        print(f"\n[REPORT] XLSX generation failed: {exc}")
        traceback.print_exc()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _infer_category(item: pytest.Item) -> str:
    for name, _ in _MARKERS:
        if item.get_closest_marker(name):
            return name
    return _FILENAME_TO_CATEGORY.get(Path(str(item.fspath)).stem, "other")


def _append_result(report: pytest.TestReport, status: str) -> None:
    parts = report.nodeid.split("::")
    _results.append({
        "nodeid":   report.nodeid,
        "name":     parts[-1],
        "module":   Path(parts[0]).stem if parts else "unknown",
        "category": _nodeid_to_category.get(report.nodeid, "other"),
        "status":   status,
        "duration": round(getattr(report, "duration", 0.0), 3),
        "error":    _short_error(report) if status in ("FAILED", "ERROR") else "",
    })


def _short_error(report: pytest.TestReport) -> str:
    if not report.longrepr:
        return ""
    try:
        text = str(report.longrepr).strip()
        lines = [ln for ln in text.split("\n") if ln.strip()]
        return lines[-1][:300] if lines else text[:300]
    except Exception:
        return "(could not extract error)"


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False


_DHL_RED    = "D40511"
_DHL_YELLOW = "FFCC00"
_WHITE      = "FFFFFF"
_DARK_HDR   = "2D2D2D"
_GREEN_FILL = "D4EDDA"
_RED_FILL   = "F8D7DA"
_SKIP_FILL  = "FFF3CD"
_ERR_FILL   = "FDE8D8"
_STRIPE     = "F8F9FB"

_STATUS_FILL = {
    "PASSED":  _GREEN_FILL,
    "FAILED":  _RED_FILL,
    "SKIPPED": _SKIP_FILL,
    "ERROR":   _ERR_FILL,
}
_STATUS_ICON = {
    "PASSED":  "PASSED",
    "FAILED":  "FAILED",
    "SKIPPED": "SKIPPED",
    "ERROR":   "ERROR",
}
_CAT_LABEL = {
    "smoke":          "Smoke Tests",
    "regression":     "Regression Tests",
    "e2e_positive":   "E2E Positive",
    "e2e_negative":   "E2E Negative",
    "api":            "API Tests",
    "integration":    "Integration Tests",
    "transformation": "Transformation Tests",
    "other":          "Other",
}
_SHEET_NAME = {
    "smoke":          "Smoke",
    "regression":     "Regression",
    "e2e_positive":   "E2E Positive",
    "e2e_negative":   "E2E Negative",
    "api":            "API",
    "integration":    "Integration",
    "transformation": "Transformation",
    "other":          "Other",
}


def _fill(hex_color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, color: str = "000000", size: int = 11) -> "Font":
    return Font(bold=bold, color=color, size=size)


def _thin_border() -> "Border":
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _write_xlsx(results: list[dict], path: Path) -> None:
    if not _XLSX_OK:
        return

    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _build_summary(ws_sum, results)

    for cat in _CATEGORY_ORDER + ["other"]:
        cat_rows = [r for r in results if r["category"] == cat]
        if not cat_rows:
            continue
        ws = wb.create_sheet(_SHEET_NAME[cat])
        _build_category_sheet(ws, _CAT_LABEL.get(cat, cat), cat_rows)

    wb.save(path)


def _build_summary(ws: Any, results: list[dict]) -> None:
    # ── Row 1: DHL red title bar ──────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "NgReact V3  —  Angular → React Migration Platform  —  Test Report"
    ws["A1"].font      = _font(bold=True, color=_WHITE, size=14)
    ws["A1"].fill      = _fill(_DHL_RED)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Row 2: yellow run-info bar ────────────────────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"] = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
        f"     |     Total executed: {len(results)}"
    )
    ws["A2"].font      = _font(size=11)
    ws["A2"].fill      = _fill(_DHL_YELLOW)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 8  # spacer

    # ── Row 4: section label ──────────────────────────────────────────────
    ws.merge_cells("A4:G4")
    ws["A4"] = "TEST CATEGORY SUMMARY"
    ws["A4"].font      = _font(bold=True, size=10)
    ws["A4"].fill      = _fill("EEEEEE")
    ws["A4"].alignment = Alignment(horizontal="center")

    # ── Row 5: column headers ─────────────────────────────────────────────
    headers = ["Category", "Total", "Passed", "Failed", "Skipped", "Error", "Pass Rate"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font      = _font(bold=True, color=_WHITE, size=10)
        cell.fill      = _fill(_DARK_HDR)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = _thin_border()
    ws.row_dimensions[5].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────
    row = 6
    totals = dict(total=0, passed=0, failed=0, skipped=0, error=0)

    for cat in _CATEGORY_ORDER + ["other"]:
        cat_rows = [r for r in results if r["category"] == cat]
        if not cat_rows:
            continue
        total   = len(cat_rows)
        passed  = sum(1 for r in cat_rows if r["status"] == "PASSED")
        failed  = sum(1 for r in cat_rows if r["status"] == "FAILED")
        skipped = sum(1 for r in cat_rows if r["status"] == "SKIPPED")
        error   = sum(1 for r in cat_rows if r["status"] == "ERROR")
        rate    = f"{passed / total * 100:.0f}%" if total else "—"

        if failed == 0 and error == 0:
            row_fill = _GREEN_FILL
        elif total and (passed / total) >= 0.8:
            row_fill = _SKIP_FILL
        else:
            row_fill = _RED_FILL

        data = [_CAT_LABEL.get(cat, cat), total, passed, failed, skipped, error, rate]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill      = _fill(row_fill)
            cell.border    = _thin_border()
            cell.alignment = Alignment(horizontal="left" if col == 1 else "center")
            if col == 1:
                cell.font = _font(bold=True)

        totals["total"]   += total
        totals["passed"]  += passed
        totals["failed"]  += failed
        totals["skipped"] += skipped
        totals["error"]   += error
        row += 1

    # ── Totals row ────────────────────────────────────────────────────────
    row += 1
    overall = (f"{totals['passed'] / totals['total'] * 100:.0f}%"
               if totals["total"] else "—")
    total_fill = _GREEN_FILL if totals["failed"] == 0 and totals["error"] == 0 else _RED_FILL
    for col, val in enumerate(
        ["TOTAL", totals["total"], totals["passed"],
         totals["failed"], totals["skipped"], totals["error"], overall], 1
    ):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = _font(bold=True, color=_WHITE)
        cell.fill      = _fill(_DHL_RED)
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="left" if col == 1 else "center")

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    for letter in "BCDEFG":
        ws.column_dimensions[letter].width = 13


def _build_category_sheet(ws: Any, title: str, results: list[dict]) -> None:
    total   = len(results)
    passed  = sum(1 for r in results if r["status"] == "PASSED")
    failed  = sum(1 for r in results if r["status"] == "FAILED")
    skipped = sum(1 for r in results if r["status"] == "SKIPPED")

    # ── Row 1: DHL red title ──────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = f"{title}  —  {total} tests"
    ws["A1"].font      = _font(bold=True, color=_WHITE, size=13)
    ws["A1"].fill      = _fill(_DHL_RED)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: yellow stats bar ───────────────────────────────────────────
    ws.merge_cells("A2:F2")
    ws["A2"] = f"PASSED: {passed}    FAILED: {failed}    SKIPPED: {skipped}"
    ws["A2"].fill      = _fill(_DHL_YELLOW)
    ws["A2"].font      = _font(size=11)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8  # spacer

    # ── Row 4: column headers ─────────────────────────────────────────────
    headers = ["#", "Test Name", "Module", "Status", "Duration (s)", "Error / Reason"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = _font(bold=True, color=_WHITE, size=10)
        cell.fill      = _fill(_DARK_HDR)
        cell.border    = _thin_border()
        cell.alignment = Alignment(horizontal="center" if col != 2 else "left")
    ws.row_dimensions[4].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────
    for i, result in enumerate(results, 1):
        r = i + 4
        status     = result["status"]
        fill_color = _STATUS_FILL.get(status, "FFFFFF")
        stripe     = _STRIPE if (i % 2 == 0 and status == "PASSED") else fill_color

        values = [
            i,
            result["name"],
            result["module"],
            _STATUS_ICON.get(status, status),
            result["duration"],
            result["error"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill      = _fill(stripe)
            cell.border    = _thin_border()
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 4, 5) else "left",
                wrap_text=(col == 6),
            )
            if col == 2 and status == "FAILED":
                cell.font = _font(bold=True)

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 58
